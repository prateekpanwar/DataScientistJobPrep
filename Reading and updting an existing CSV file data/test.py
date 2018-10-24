import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('Walmart Inventory Template.xlsx')
#writer = pd.ExcelWriter('Walmart Inventory Template.xlsx', engine='openpyxl') 
#writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#walmart.iloc[:,3].to_excel(writer, sheet_name='Sheet1', startrow=5, startcol=3, index=False, header=False, encoding='utf8')
#data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

#writer.write_cells

#writer.save()
#writer.close()

ws = wb.get_active_sheet()
ws.cell(row=6,column=4).value = 55
wb.save('Walmart Inventory Template.xlsx')
wb.close()
