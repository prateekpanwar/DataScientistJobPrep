# -*- coding: utf-8 -*-
"""
Created on Sat Oct 13 22:49:14 2018

@author: hp1
"""

import pandas as pd
from openpyxl import load_workbook

walmart = pd.read_excel('Walmart Inventory Template.xlsx', skiprows=4)

inventory = pd.read_excel('Inventory Master Oct 10 2018 - TEST.xlsm', sheet_name='Total Quantity')

count = 0
for sku in inventory['SKU']:
    if sku in list(walmart['your sku #']):
        walmart.iloc[list(walmart['your sku #']).index(sku), 3] = inventory.iloc[count, 1]
    else:
        print('no')
    count=count+1
              
wb = load_workbook('Walmart Inventory Template.xlsx')
ws = wb.get_active_sheet()

for i in range(len(walmart.iloc[:,3])):
    ws.cell(row=i+6,column=4).value = walmart.iloc[i, 3]

wb.save('Walmart Inventory Template.xlsx')
wb.close()